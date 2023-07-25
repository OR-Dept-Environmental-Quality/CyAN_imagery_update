# Import of CyAN data from NASA

# Scripts to download and process daily CyAN data (EPA) from NASA

# Originally developed by Brian Fulfrost in 10/2020

# Modified by Dan Sobota in 5/2021; Yuan Grund by 1/5/2022

# Script functions

# This script downloads NASA OLCI imagery from NASA's website, calculates cyanobacteria abundance according to
# the US EPA CyAN project protocol, and calculates summary statistics for the 42 resolvable lakes/reservoirs in
# Oregon. See specific code chunks for the explanations of the code functions.
# This code is specific to Sentinel 3 data from 2016 to the present.

# Requirments:

# This script is run within the arcgispro-py3 virtual environment; you must have access to arcpy to run the scripts
# A Spatial Analyst license is needed to run the zonal statistics feature

# Import libraries

# from CyANconfig import *
# Configuration file with file paths and website addresses; place in same folder as this python script
import sys
import urllib.request
import ssl
import shutil
import os
import datetime
import pandas as pd
import glob
from openpyxl import load_workbook
from simpledbf import Dbf5
import arcpy
from arcpy import env
from arcpy.sa import *
from datetime import date

# Check to make sure right version of python is in use (3.7.10)
print(sys.version_info)
if sys.version_info < (3, 7):
    print('Please upgrade your Python version to 3.7.0 or higher')
    sys.exit()

# Define the base URL and check that it is a valid address
# baseurl = "https://oceancolor.gsfc.nasa.gov/CYAN/OLCI/"

try:
    ssl._create_default_https_context = ssl._create_unverified_context
    urllib.request.urlopen(baseurl)
    print("Valid url:", baseurl)
except:
    print("Invalid url")

# Gets the appropriate date range for the current query
# Note that the query can only be done for the current year
year = str(date.today().year)
# year = str("2022")
print("Year:", year)

# Get the start date
# tif_file_path = '\\\\deqhq1\\wq-share\\Harmful Algal Blooms Coordination Team\\HAB_Shiny_app\\data\\2023\\*.tif'
list_of_files = glob.glob(tif_file_path)  # Change directory as needed
latest_file = max(list_of_files, key=os.path.getctime)  # Getting the last day from the last update
hab_day_start = int(latest_file[85:88]) + 1  # assumes weekly updates; adjust as needed
# hab_day_start = int(1) used when a new year starts, and "hab_day_start = int(latest_file[85:88]) + 1" need to be muted
# hab_day_start = int(1)
# hab_day_start = 297
print("start day:", hab_day_start)

hab_day_end = date.today().timetuple().tm_yday - 1  # assumes data from the previous day available
# hab_day_end = 343
# hab_day_end = int(365)
print("end day:", hab_day_end)

# Generate sequence
hab_days = list(range(hab_day_start, hab_day_end + 1))
hab_days = [str(i).zfill(3) for i in hab_days]  # convert to strings with leading zeros
hab_days_length = len(hab_days)
print("day sequence:", hab_days)
print("sequence length:", hab_days_length)

# Set archive and filename variables for date range
# file = .tgz archive

file = ['L' + year + hab_days[i] + '.L3m_DAY_CYAN_CI_cyano_CYAN_CONUS_300m.tgz'
        for i in range(0, hab_days_length)]
file_name = ['L' + year + hab_days[i] + '.L3m_DAY_CYAN_CI_cyano_CYAN_CONUS_300m.tgz'
             for i in range(0, hab_days_length)]
file2 = ['L' + year + hab_days[i] + '.L3m_DAY_CYAN_CI_cyano_CYAN_CONUS_300m'
         for i in range(0, hab_days_length)]
url = [baseurl + year + '/' + hab_days[i] + '/' + file[i]
       for i in range(0, hab_days_length)]
# extract_base = '\\\\deqhq1\\wq-share\\Harmful Algal Blooms Coordination Team\\GIS\\cyan\\'
extract_path = extract_base + year  # Change as needed

# Step 2c - download, extract, and rename imagery for date range
for i in range(0, hab_days_length):
  # test:   i = 6
    print(url[i])

    # BKF - add some code that doesnt download if .tgz file already exist in _tgz subdirectory
    # BKF - also need to add some timeout code in case of connection issues
    urllib.request.urlretrieve(url[i], extract_path + file[i])
    shutil.unpack_archive(extract_path + file[i], extract_path)

    # BKF - need to point to a DEQ server directory (for universal application)
    shutil.move(extract_path + file[i], extract_path + '\\_tif\\' + file[i])

# Need to make directories for moving Oregon files

archive_dir = [os.path.join(extract_path, file2[m])
               for m in range(0, hab_days_length)]

for n in range(0, hab_days_length):
    os.mkdir(os.path.join(archive_dir[n], 'temp'))

temp_dir = [os.path.join(archive_dir[o], 'temp')
            for o in range(0, hab_days_length)]

# Copy Oregon images to temp subdirectory
# Oregon includes: 1_1, 1_2, 2_1, 2_2

for i in range(0, hab_days_length):

    # BKF - copy image files for oregon to temp directory
    # BKF - this will be useful to use directly one NASA starts to post individual files instead of .tgz
    shutil.copyfile(os.path.join(archive_dir[i], file2[i] + '_1_1.tif'),
                    os.path.join(temp_dir[i], file2[i] + '_1_1.tif'))
    shutil.copyfile(os.path.join(archive_dir[i], file2[i] + '_1_2.tif'),
                    os.path.join(temp_dir[i], file2[i] + '_1_2.tif'))
    shutil.copyfile(os.path.join(archive_dir[i], file2[i] + '_2_1.tif'),
                    os.path.join(temp_dir[i], file2[i] + '_2_1.tif'))
    shutil.copyfile(os.path.join(archive_dir[i], file2[i] + '_2_2.tif'),
                    os.path.join(temp_dir[i], file2[i] + '_2_2.tif'))

    # rename oregon images
    env.workspace = temp_dir[i]
    arcpy.CheckOutExtension("Spatial")

    for raster in arcpy.ListRasters():
        fileName, fileExtension = os.path.splitext(raster)
        tile = fileName[-4:]

        fileNameParts = fileName.split('.')
        compactFileName = fileNameParts[0] + tile + fileExtension
        print(compactFileName)

        arcpy.Rename_management(raster, compactFileName)
        
print("done")

# Convert to cells/ml , mosaic 4 tiles into one oregon image, calc zonal stats for resolvable lakes
# Setnull and cellsml

for i in range(0, hab_days_length):
  
    # Test: 
    i=5
    # rename oregon images
    env.workspace = temp_dir[i]
    arcpy.CheckOutExtension("Spatial")
    os.mkdir(os.path.join(temp_dir[i], 'cellsml'))
    cellsml_dir = os.path.join(temp_dir[i], 'cellsml')
    output_DIR = os.path.join('c:', 'hab', 'cyan', year, 'cellsml')
    for raster in arcpy.ListRasters():
        # Test: 
        raster = "L2023203_2_2.tif"
        # set non cyano values to null
        outSetNull = SetNull(raster, raster, 'VALUE = 255 or VALUE = 254')
        # 254 = land; 255 = water
        # convert CyAN index numbers to cells/ml
        outCellsML = (Power(10, (3.0 / 250.0 * Int(outSetNull) - 4.2))) * 100000000

        # save the cells/ml to subdirectory
        final_DIR = os.path.join(cellsml_dir, raster)
        outCellsML.save(final_DIR)

        del outSetNull
        del outCellsML
        
print("done")

# Make mosaic directory
# os.mkdir(os.path.join(extract_path, 'mosaic')) # Needed for start of new year
mosaicdir = os.path.join(extract_path, 'mosaic')

# Mosaic
for i in range(0, hab_days_length):
    cellsml_dir = os.path.join(extract_path,temp_dir[i], 'cellsml')
    env.workspace = cellsml_dir

    sr = arcpy.SpatialReference()
    sr.factoryCode = 5070
    sr.create()

    mosaicdict = {}

    for raster in arcpy.ListRasters():
        fileName, fileExtension = os.path.splitext(raster)
        mosaickey = fileName[1:8]
        mosaicfilename = fileName[1:8] + ".tif"
        if mosaickey not in mosaicdict:
            mosaicdict[mosaickey] = []

        if len(mosaicdict[mosaickey]) == 0:
            mosaicdict[mosaickey].append(raster)
        else:
            mosaicdict[mosaickey].append(raster)
            if len(mosaicdict[mosaickey]) == 4:
                print(mosaicfilename)
                arcpy.MosaicToNewRaster_management(mosaicdict[mosaickey], mosaicdir, mosaicfilename, sr,
                                                   "32_BIT_FLOAT", "300", "1", "LAST", "FIRST")
print("done")

# Need to get a list of file names for zonal statistics
# Need to modify to account for existing files in the directory

mosaicfilename2 = list()
for i, file in enumerate(os.listdir(mosaicdir)):
    if file.endswith(".tif"):
        mosaicfilename2.append(os.path.basename(file))

mosaicfilename2 = mosaicfilename2[(hab_day_start - 1):(hab_day_end + 1)]

# Need to get a list of names for mosaic key
# Need to modify to account for existing files in the directory

mosaickey2 = list()
for i, file in enumerate(os.listdir(mosaicdir)):
    if file.endswith(".tif"):
        mosaickey2.append(os.path.splitext(file)[0])

mosaickey2 = mosaickey2[(hab_day_start - 1):(hab_day_end + 1)]

# Zonal stats
for i in range(0, hab_days_length):
  #test:        
  i=6
    zonalraster = os.path.join(mosaicdir, mosaicfilename2[i])
    # zones = r"\\deqhq1\wq-share\Harmful Algal Blooms Coordination Team\GIS\cyan\HAB_deschutes2020.gdb\NHDWaterbody_resolvable_lakes"
    stats_dir = os.path.join(extract_path, 'mosaic', 'stats')

    thestatsname = os.path.join(os.path.join(stats_dir, mosaickey2[i] + "_stats.dbf"))
    print(thestatsname)

    # Process: Zonal Statistics as Table
    env.workspace = mosaicdir
    arcpy.gp.ZonalStatisticsAsTable_sa(zones, "GNISIDNAME", zonalraster, thestatsname, "DATA", "ALL")
   
    arcpy.DeleteField_management(thestatsname, ["ZONE_CODE", "SUM", "MEDIAN", "PCT90"])

    # calcdaydate
    arcpy.AddField_management(thestatsname, 'Day', 'LONG')
    arcpy.AddField_management(thestatsname, 'Year', 'LONG')
    arcpy.AddField_management(thestatsname, 'Date', 'Date')

    # extract theyear and theday
    yeardate = mosaickey2[i]
    theyear = int(yeardate[0:4])
    theday = int(yeardate[4:7])

    # used when working with individual files instead of mosaics - includes 'L'
    # theyear = int(fileName[1:5])
    # theday = int(fileName[5:8])

    # calc day and year
    arcpy.CalculateField_management(thestatsname, 'Day', theday, 'PYTHON')
    arcpy.CalculateField_management(thestatsname, 'Year', theyear, 'PYTHON')

    # calc date
    thedate = datetime.datetime(theyear, 1, 1) + datetime.timedelta(theday - 1)
    # newdate = thedate.strftime('%Y-%m-%d')
    epoch = datetime.datetime(1899, 12, 30)
    days = (thedate - epoch).days
    arcpy.CalculateField_management(thestatsname, 'Date', days, 'PYTHON')

    # deleted temp rasters above and directory removal below now works
    shutil.rmtree(os.path.join(temp_dir[i], 'cellsml'), ignore_errors=True)
    shutil.rmtree(temp_dir[i], ignore_errors=True)
    shutil.rmtree(archive_dir[i], ignore_errors=True)
    
print("done")

# Need to get a list of stat files for mosaic key
stats_dir = os.path.join(extract_path, 'mosaic', 'stats')

thestatsname = list()
for i, file in enumerate(os.listdir(stats_dir)):
    if file.endswith(".dbf"):
        thestatsname.append(os.path.basename(file))

thestatsname = thestatsname[(hab_day_start - 2):(hab_day_end + 1)]

# Convert table to df, rename fields, add new field, and append excel file
# for i in range(0, hab_days_length):
#     # test: i = 0
#     dbf = Dbf5(os.path.join(stats_dir, thestatsname[i]))
#     df = dbf.to_dataframe()
# 
#     # add new field and rename existing fields
#     df.insert(3, "PercentArea_Value", '', True)
#     df.rename(columns={"MIN": "MIN_cellsml", "MAX": "MAX_cellsml", "RANGE": "RANGE_cellsml", "MEAN": "MEAN_cellsml",
#                        "STD": "STD_cellsml"})
# 
#     # append new data to exsiting excel spreadhseet
#     # BKK - spreadhseet and worksheet shouldn't have date
#     # dir_Shiny = "\\\\deqhq1\\wq-share\\Harmful Algal Blooms Coordination Team\\HAB_Shiny_app\\data"
#     thetable = os.path.join(dir_Shiny, 'HAB_resolvablelakes_2022.xlsx')
#     writer = pd.ExcelWriter(thetable, engine='openpyxl', mode='a', if_sheet_exists='overlay')
#     writer.book = load_workbook(thetable)
#     writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
#     # BKF - need to make sure that startrow isn't overwriting last row
#     firstrow = writer.book['HAB_resolvable_lake_data'].max_row
# 
#     df.to_excel(writer, sheet_name='HAB_resolvable_lake_data', startrow=firstrow, startcol=0, index=False, header=None)
# 
#     writer.save()
#     #writer.close

# Set workspace and directory for reprojection; manually adjust for now
env.workspace = os.path.join(extract_path, "mosaic")
arcpy.env.compression = "LZW"
arcpy.env.overwriteOutput = True
final_dir = os.path.join(extract_path, "mosaic", "web")
final_dir_Shiny = os.path.join(dir_Shiny, year)

# Set up spatial data for reprojection

arcpy.CheckOutExtension("Spatial")

# themask = r"\\deqhq1\wq-share\Harmful Algal Blooms Coordination Team\GIS\cyan\HAB_deschutes2020.gdb\stateline_buffer50albers"
# theextent = r"\\deqhq1\wq-share\Harmful Algal Blooms Coordination Team\GIS\cyan\HAB_deschutes2020.gdb\stateline_buffer50web"

for raster in mosaicfilename2:
  # test:    raster = "2023169.tif"
    fileName, fileExtension = os.path.splitext(raster)
    tile = fileName[-8:]
    tmpfile = "temp.tif"
    newfile = (tile + fileExtension)
    print(os.path.join(final_dir_Shiny, tile + fileExtension))

    arcpy.env.extent = theextent
    arcpy.ProjectRaster_management(in_raster=raster, out_raster=os.path.join(final_dir_Shiny, tmpfile),
                                   out_coor_system="PROJCS['WGS_1984_Web_Mercator_Auxiliary_Sphere',GEOGCS['GCS_WGS_1984',DATUM['D_WGS_1984',SPHEROID['WGS_1984',6378137.0,298.257223563]],PRIMEM['Greenwich',0.0],UNIT['Degree',0.0174532925199433]],PROJECTION['Mercator_Auxiliary_Sphere'],PARAMETER['False_Easting',0.0],PARAMETER['False_Northing',0.0],PARAMETER['Central_Meridian',0.0],PARAMETER['Standard_Parallel_1',0.0],PARAMETER['Auxiliary_Sphere_Type',0.0],UNIT['Meter',1.0]]",
                                   resampling_type="NEAREST", cell_size="300 300",
                                   geographic_transform="WGS_1984_(ITRF00)_To_NAD_1983", Registration_Point="",
                                   in_coor_system="PROJCS['USA_Contiguous_Albers_Equal_Area_Conic_USGS_version',GEOGCS['GCS_North_American_1983',DATUM['D_North_American_1983',SPHEROID['GRS_1980',6378137.0,298.257222101]],PRIMEM['Greenwich',0.0],UNIT['Degree',0.0174532925199433]],PROJECTION['Albers'],PARAMETER['False_Easting',0.0],PARAMETER['False_Northing',0.0],PARAMETER['Central_Meridian',-96.0],PARAMETER['Standard_Parallel_1',29.5],PARAMETER['Standard_Parallel_2',45.5],PARAMETER['Latitude_Of_Origin',23.0],UNIT['Meter',1.0]]")
    tmpraster = os.path.join(final_dir_Shiny, tmpfile)
    finalraster = os.path.join(final_dir_Shiny, newfile)

    arcpy.Clip_management(in_raster=tmpraster,
                          rectangle="-13975366.498500 5052033.819600 -12850574.046900 5935465.862100",
                          out_raster=finalraster,
                          in_template_dataset=themask, nodata_value="-3.402823e+038",
                          clipping_geometry="ClippingGeometry", maintain_clipping_extent="NO_MAINTAIN_EXTENT")

print("done")

